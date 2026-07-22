#!/usr/bin/env python3
"""T3b Phase 0 — full export of Canonical_DB.xlsx import families + normalization config → CSV.

Read-only against the handoff package. Emits:
  .tmp/handoff-full/families/*.csv   (the 6 canonical-input families, full rows)
  .tmp/handoff-full/_norm/*.csv      (normalization layer: crosswalk / predicate map / vocab / flags)

Idempotent: re-running overwrites the CSVs. Prints a row-count summary for the manifest.
"""
import csv
import sys
from pathlib import Path

import openpyxl

SRC = Path("docs/RKC_Handoff_July_2026_FINAL_VERIFIED/02_Core/Canonical_DB.xlsx")
OUT = Path(".tmp/handoff-full")

# sheet name -> output filename (family sheets)
FAMILIES = {
    "Source-System Cards": "source-system-cards.csv",
    "New Objects": "new-objects.csv",
    "Option Candidates": "option-candidates.csv",
    "Claims and Cautions": "claims-cautions.csv",
    "Implementation Memory": "implementation-memory.csv",
    "Relationship Leads": "relationship-leads.csv",
}

# sheet name -> output filename (normalization config)
NORM = {
    "Object Type Crosswalk": "object-type-crosswalk.csv",
    "Relationship Predicate Map": "predicate-map.csv",
    "Controlled Vocabularies": "controlled-vocab.csv",
    "Normalization Flags": "normalization-flags.csv",
    "Option Normalization Index": "option-normalization-index.csv",
    "Claim-Caution Normalization": "claim-caution-normalization.csv",
    "Operational Normalization": "operational-normalization.csv",
}


def export_sheet(ws, dest: Path) -> int:
    """Write a worksheet to CSV. Returns data-row count (excludes header)."""
    dest.parent.mkdir(parents=True, exist_ok=True)
    rows = list(ws.iter_rows(values_only=True))
    # drop fully-empty leading rows
    while rows and not any(c is not None and str(c).strip() for c in rows[0]):
        rows.pop(0)
    if not rows:
        dest.write_text("")
        return 0
    with dest.open("w", newline="", encoding="utf-8") as f:
        w = csv.writer(f)
        for row in rows:
            w.writerow(["" if c is None else c for c in row])
    # data rows = total minus header, minus trailing empties
    data = [r for r in rows[1:] if any(c is not None and str(c).strip() for c in r)]
    return len(data)


def main():
    if not SRC.exists():
        sys.exit(f"source not found: {SRC}")
    wb = openpyxl.load_workbook(SRC, read_only=True, data_only=True)
    present = set(wb.sheetnames)

    print("=== FAMILIES (canonical input) ===")
    fam_total = 0
    for sheet, fname in FAMILIES.items():
        if sheet not in present:
            print(f"  MISSING sheet: {sheet}")
            continue
        n = export_sheet(wb[sheet], OUT / "families" / fname)
        fam_total += n
        print(f"  {sheet:26s} -> families/{fname:28s} {n:>5} rows")
    print(f"  {'TOTAL':26s} {'':40s} {fam_total:>5} rows")

    print("\n=== NORMALIZATION CONFIG ===")
    for sheet, fname in NORM.items():
        if sheet not in present:
            print(f"  MISSING sheet: {sheet}")
            continue
        n = export_sheet(wb[sheet], OUT / "_norm" / fname)
        print(f"  {sheet:30s} -> _norm/{fname:34s} {n:>5} rows")


if __name__ == "__main__":
    main()
